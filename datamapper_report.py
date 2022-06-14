import os, sys
from openpyxl import Workbook

sheet = Workbook()
filepath = 'report.xlsx'
ws = sheet.active

ws['A1'] = "Nr"
ws['B1'] = "Link nr"
ws['C1'] = "Destination"
ws['D1'] = "Link type"
ws['E1'] = "Condition"
ws['F1'] = "Triggering"
ws['G1'] = "FIRST"
ws['H1'] = "EACH"
ws['I1'] = "LAST"
ws['J1'] = "Expression"
ws['K1'] = "Source"

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 20
ws.column_dimensions['J'].width = 20
ws.column_dimensions['K'].width = 20


f = open("T_MAP_PL_DANFTSO4IS.rep", "r")
i = 1
first = []
each = []
last = []

for index, line in enumerate(f):
    if "===D=" in line:
        i += 1
        ws.cell(row=i, column=1).value = index + 1
        ws.cell(row=i, column=2).value = line[line.find('=[')+2 : line.find(']=')]
        ws.cell(row=i, column=3).value = line[line.find('>>')+3 : len(line)]
        ws.cell(row=i, column=4).value = "Driver"
        ws.cell(row=i, column=6).value = line[line.find('=(')+2 : line.find(')=')]
    elif "===C=" in line:
        i += 1
        ws.cell(row=i, column=1).value = index + 1
        ws.cell(row=i, column=2).value = line[line.find('=[')+2 : line.find(']=')]
        ws.cell(row=i, column=3).value = line[line.find('/>')+3 : len(line)]
        ws.cell(row=i, column=4).value = "Contributor"
        ws.cell(row=i, column=6).value = line[line.find('=(') + 2: line.find(')=')]
    elif "---->" in line:
        i += 1
        ws.cell(row=i, column=1).value = index + 1
        ws.cell(row=i, column=2).value = line[line.find('-[')+2 : line.find(']-')]
        ws.cell(row=i, column=3).value = line[line.find('->')+3 : len(line)]
        ws.cell(row=i, column=4).value = "data"

    if "DLN Cond" in line:
        ws.cell(row=i, column=5).value = line[line.find('Cond')+5 : len(line)]

    if "|	First" in line:
        first.append(line[line.find(':')+2 : len(line)])
    else: #this is the first line when First is no more, so we need to append it to excel:
        element = ''
        for _ in first:
            element = element + _
        if element != '':
            ws.cell(row=i, column=7).value = element   #FORMATIING TBD
        first = []






sheet.save(filepath)
f.close()